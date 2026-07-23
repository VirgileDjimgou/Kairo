"""Import ordinary members and 2026 contributions from a controlled workbook."""

from __future__ import annotations

import argparse
import asyncio
import csv
import json
import logging
import re
import secrets
import unicodedata
import uuid
from dataclasses import dataclass
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import delete, select

from app.core.security import hash_password
from app.db.session import async_session_factory
from app.modules.contributions.models import ContributionRecord, ContributionReminder, PaymentRecord
from app.modules.disciplinary.models import DisciplinaryRecord
from app.modules.identity.models import User
from app.modules.membership.models import MembershipProfile
from app.modules.tenancy.models import Role, Tenant, TenantUser, user_roles

SHEET_NAME = "Feuille 2"
PAYMENT_COLUMNS = (
    "inscription",
    "1ere tranche",
    "2em tranche",
    "3em tranche",
    "4em tranche",
    "5em tranche",
    "6em tranche",
)


@dataclass(frozen=True)
class MemberImportRow:
    sequence: int
    first_name: str
    last_name: str
    remark: str
    payments: dict[str, Decimal]
    total: Decimal

    @property
    def display_name(self) -> str:
        return f"{self.first_name} {self.last_name}".strip()

    @property
    def login_email(self) -> str:
        return f"{self.sequence:04d}-{slugify(self.last_name)}@demo.local"

    @property
    def member_code(self) -> str:
        return f"IMP-{self.sequence:04d}"


def slugify(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_value = normalized.encode("ascii", "ignore").decode("ascii").lower()
    slug = re.sub(r"[^a-z0-9]+", "-", ascii_value).strip("-")
    return slug or "member"


def parse_amount(value: object) -> Decimal:
    if value in (None, ""):
        return Decimal("0.00")
    try:
        amount = Decimal(str(value).replace(",", "."))
    except (InvalidOperation, ValueError) as error:
        raise ValueError(f"Invalid monetary value: {value!r}") from error
    if amount < 0:
        raise ValueError("Monetary values must not be negative.")
    return amount.quantize(Decimal("0.01"))


def value_at(values: tuple[object, ...], index: int) -> object:
    return values[index] if index < len(values) else None


def load_rows(workbook_path: Path, sheet_name: str) -> list[MemberImportRow]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Worksheet not found: {sheet_name}")

    worksheet = workbook[sheet_name]
    source_rows = worksheet.iter_rows(values_only=True)
    header = next(source_rows, None)
    if header is None:
        raise ValueError("The worksheet is empty.")
    headers = [str(value).strip().lower() if value is not None else "" for value in header]
    expected_headers = {"nom", "prenom", "total", *PAYMENT_COLUMNS}
    missing_headers = expected_headers.difference(headers)
    if missing_headers:
        raise ValueError(f"Missing required columns: {', '.join(sorted(missing_headers))}")

    column_index = {name: index for index, name in enumerate(headers)}
    imported_rows: list[MemberImportRow] = []
    for values in source_rows:
        if not any(value not in (None, "") for value in values):
            continue
        last_name = str(value_at(values, column_index["nom"]) or "").strip()
        if not last_name:
            raise ValueError("Every imported member must have a last name.")
        first_name = str(value_at(values, column_index["prenom"]) or "Membre").strip() or "Membre"
        payments = {
            column: parse_amount(value_at(values, column_index[column]))
            for column in PAYMENT_COLUMNS
        }
        source_total = parse_amount(value_at(values, column_index["total"]))
        calculated_total = sum(payments.values(), Decimal("0.00"))
        total = source_total if source_total else calculated_total
        remark_index = column_index.get("remarque")
        remark = ""
        if remark_index is not None:
            remark = str(value_at(values, remark_index) or "").strip()
        imported_rows.append(
            MemberImportRow(
                sequence=len(imported_rows) + 1,
                first_name=first_name,
                last_name=last_name,
                remark=remark,
                payments=payments,
                total=total,
            )
        )

    if not imported_rows:
        raise ValueError("No member rows were found in the worksheet.")
    if len(imported_rows) > 9999:
        raise ValueError("The identifier format supports at most 9,999 members.")
    return imported_rows


async def find_ordinary_member_users(session, tenant_id: uuid.UUID) -> list[TenantUser]:
    tenant_users = (
        await session.execute(select(TenantUser).where(TenantUser.tenant_id == tenant_id))
    ).scalars().all()
    ordinary_members: list[TenantUser] = []
    for tenant_user in tenant_users:
        role_codes = set(
            (
                await session.execute(
                    select(Role.code)
                    .select_from(user_roles)
                    .join(Role, Role.id == user_roles.c.role_id)
                    .where(user_roles.c.tenant_user_id == tenant_user.id)
                )
            ).scalars()
        )
        if role_codes == {"member"}:
            ordinary_members.append(tenant_user)
    return ordinary_members


async def import_members(
    workbook_path: Path,
    credentials_output: Path,
    tenant_slug: str,
    year: int,
    execute: bool,
) -> tuple[int, int]:
    rows = load_rows(workbook_path, SHEET_NAME)
    if not execute:
        return len(rows), 0

    credentials: list[dict[str, str]] = []
    credentials_output.parent.mkdir(parents=True, exist_ok=True)
    temporary_credentials = credentials_output.with_suffix(".tmp")

    async with async_session_factory() as session:
        async with session.begin():
            tenant = (
                await session.execute(select(Tenant).where(Tenant.slug == tenant_slug))
            ).scalar_one_or_none()
            if tenant is None:
                raise ValueError(f"Tenant not found: {tenant_slug}")
            member_role = (
                await session.execute(
                    select(Role).where(Role.tenant_id == tenant.id, Role.code == "member")
                )
            ).scalar_one()
            ordinary_members = await find_ordinary_member_users(session, tenant.id)
            ordinary_user_ids = [tenant_user.user_id for tenant_user in ordinary_members]
            ordinary_tenant_user_ids = [tenant_user.id for tenant_user in ordinary_members]
            profile_ids = list(
                (
                    await session.execute(
                        select(MembershipProfile.id).where(
                            MembershipProfile.tenant_id == tenant.id,
                            MembershipProfile.user_id.in_(ordinary_user_ids),
                        )
                    )
                ).scalars()
            ) if ordinary_user_ids else []

            if profile_ids:
                contribution_ids = list(
                    (
                        await session.execute(
                            select(ContributionRecord.id).where(
                                ContributionRecord.tenant_id == tenant.id,
                                ContributionRecord.membership_profile_id.in_(profile_ids),
                            )
                        )
                    ).scalars()
                )
                if contribution_ids:
                    await session.execute(
                        delete(PaymentRecord).where(
                            PaymentRecord.contribution_record_id.in_(contribution_ids)
                        )
                    )
                await session.execute(
                    delete(ContributionReminder).where(
                        ContributionReminder.membership_profile_id.in_(profile_ids)
                    )
                )
                await session.execute(
                    delete(DisciplinaryRecord).where(
                        DisciplinaryRecord.membership_profile_id.in_(profile_ids)
                    )
                )
                await session.execute(
                    delete(ContributionRecord).where(
                        ContributionRecord.membership_profile_id.in_(profile_ids)
                    )
                )
                await session.execute(
                    delete(MembershipProfile).where(MembershipProfile.id.in_(profile_ids))
                )

            if ordinary_tenant_user_ids:
                await session.execute(
                    delete(user_roles).where(user_roles.c.tenant_user_id.in_(ordinary_tenant_user_ids))
                )
                await session.execute(
                    delete(TenantUser).where(TenantUser.id.in_(ordinary_tenant_user_ids))
                )

            await session.flush()
            for user_id in ordinary_user_ids:
                other_tenant_membership = (
                    await session.execute(
                        select(TenantUser.id).where(TenantUser.user_id == user_id).limit(1)
                    )
                ).scalar_one_or_none()
                if other_tenant_membership is None:
                    await session.execute(delete(User).where(User.id == user_id))

            for row in rows:
                password = secrets.token_urlsafe(18)
                user = User(
                    id=uuid.uuid4(),
                    email=row.login_email,
                    password_hash=hash_password(password),
                    display_name=row.display_name,
                    preferred_language="fr",
                    status="active",
                )
                session.add(user)
                await session.flush()
                tenant_user = TenantUser(
                    id=uuid.uuid4(),
                    tenant_id=tenant.id,
                    user_id=user.id,
                    profile_type="member",
                    membership_status="active",
                )
                session.add(tenant_user)
                await session.flush()
                await session.execute(
                    user_roles.insert().values(
                        tenant_user_id=tenant_user.id,
                        role_id=member_role.id,
                    )
                )
                profile = MembershipProfile(
                    id=uuid.uuid4(),
                    tenant_id=tenant.id,
                    user_id=user.id,
                    member_code=row.member_code,
                    first_name=row.first_name,
                    last_name=row.last_name,
                    display_name=row.display_name,
                    email=row.login_email,
                    status="active",
                    metadata_json=json.dumps({"import_remark": row.remark}, ensure_ascii=False),
                )
                session.add(profile)
                await session.flush()
                contribution = ContributionRecord(
                    id=uuid.uuid4(),
                    tenant_id=tenant.id,
                    membership_profile_id=profile.id,
                    year=year,
                    expected_amount=row.total,
                    paid_amount=row.total,
                    balance=Decimal("0.00"),
                    currency="EUR",
                    status="paid" if row.total else "pending",
                    metadata_json=json.dumps(
                        {"source": "Feuille 2", "imported_at": datetime.now(UTC).isoformat()}
                    ),
                )
                session.add(contribution)
                await session.flush()
                for payment_label, amount in row.payments.items():
                    if amount == 0:
                        continue
                    session.add(
                        PaymentRecord(
                            id=uuid.uuid4(),
                            tenant_id=tenant.id,
                            contribution_record_id=contribution.id,
                            amount=amount,
                            currency="EUR",
                            paid_at=datetime.now(UTC),
                            payment_method="other",
                            reference=f"IMPORT-{year}-{row.sequence:04d}-{slugify(payment_label)}",
                            metadata_json=json.dumps({"source_column": payment_label}),
                        )
                    )
                credentials.append(
                    {
                        "member_code": row.member_code,
                        "display_name": row.display_name,
                        "login": row.login_email,
                        "password": password,
                    }
                )

            with temporary_credentials.open("w", newline="", encoding="utf-8") as output:
                writer = csv.DictWriter(
                    output,
                    fieldnames=["member_code", "display_name", "login", "password"],
                )
                writer.writeheader()
                writer.writerows(credentials)

    temporary_credentials.replace(credentials_output)
    return len(rows), len(ordinary_members)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--credentials-output", required=True, type=Path)
    parser.add_argument("--tenant-slug", default="demo")
    parser.add_argument("--year", type=int, default=2026)
    parser.add_argument("--execute", action="store_true")
    return parser.parse_args()


async def main() -> None:
    # Imports handle member data; never emit SQL parameter values to operator logs.
    logging.getLogger("sqlalchemy.engine").setLevel(logging.WARNING)
    logging.getLogger("sqlalchemy.engine.Engine").setLevel(logging.WARNING)
    args = parse_args()
    imported_count, replaced_count = await import_members(
        workbook_path=args.workbook,
        credentials_output=args.credentials_output,
        tenant_slug=args.tenant_slug,
        year=args.year,
        execute=args.execute,
    )
    if args.execute:
        print(f"Imported ordinary members: {imported_count}")
        print(f"Replaced ordinary member accounts: {replaced_count}")
        print(f"Credentials file: {args.credentials_output}")
    else:
        print(f"Dry run validated member rows: {imported_count}")


if __name__ == "__main__":
    asyncio.run(main())
